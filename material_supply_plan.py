import pandas as pd
import numpy as np
import re
from pathlib import Path


# =========================================================
# 0. 기본 경로 설정
# =========================================================

BASE_DIR = Path("/kaggle/working")

production_plan_path = BASE_DIR / "production_plan.xlsx"
bom_master_path = BASE_DIR / "bom_master.xlsx"
inventory_path = BASE_DIR / "inventory.xlsx"
po_open_path = BASE_DIR / "po_open.xlsx"
material_master_path = BASE_DIR / "material_master.xlsx"

output_path = BASE_DIR / "material_supply_plan.xlsx"


# =========================================================
# 1. 공통 함수 정의
# =========================================================

def print_df_info(name, df):
    """
    데이터프레임의 행 수와 컬럼명을 출력하는 함수
    """
    print(f"\n[{name}]")
    print(f"- 행 수: {len(df):,}")
    print(f"- 컬럼: {list(df.columns)}")


def check_file_exists(file_path, file_name):
    """
    파일 존재 여부 검증 함수
    """
    if not file_path.exists():
        raise FileNotFoundError(f"[{file_name}] 파일이 존재하지 않습니다: {file_path}")


def check_required_columns(df, required_columns, file_name):
    """
    필수 컬럼 존재 여부 검증 함수
    """
    missing_cols = [col for col in required_columns if col not in df.columns]

    if missing_cols:
        raise ValueError(f"[{file_name}]에 필수 컬럼이 없습니다: {missing_cols}")


def convert_numeric(df, columns):
    """
    숫자 컬럼을 숫자형으로 변환하는 함수
    변환 불가 값과 빈값은 0으로 처리
    """
    for col in columns:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    return df


def clean_columns(df):
    """
    컬럼명 앞뒤 공백 제거
    """
    df.columns = df.columns.astype(str).str.strip()
    return df


def clean_code_column(df, column_name):
    """
    코드 컬럼을 문자열로 변환하고 앞뒤 공백 제거
    """
    if column_name in df.columns:
        df[column_name] = df[column_name].astype(str).str.strip()
    return df


def parse_iso_week(plan_week_value):
    """
    plan_week 값을 날짜 정보로 변환하는 함수

    처리 가능 예시:
    - 2026-W27
    - 2026W27
    - 2026-07-05 같은 일반 날짜

    반환 컬럼:
    - week_start_date: 해당 주 월요일
    - required_date: 해당 주 일요일
    - plan_month: required_date 기준 YYYY-MM
    - month_end_date: required_date가 속한 월의 말일
    """

    value = str(plan_week_value).strip()

    match = re.match(r"^(\d{4})-?W(\d{1,2})$", value)

    if match:
        year = int(match.group(1))
        week = int(match.group(2))

        week_start_date = pd.to_datetime(
            f"{year}-W{week:02d}-1",
            format="%G-W%V-%u",
            errors="coerce"
        )

        required_date = pd.to_datetime(
            f"{year}-W{week:02d}-7",
            format="%G-W%V-%u",
            errors="coerce"
        )

        if pd.isna(required_date):
            return pd.Series({
                "week_start_date": pd.NaT,
                "required_date": pd.NaT,
                "plan_month": value,
                "month_end_date": pd.NaT
            })

        plan_month = required_date.strftime("%Y-%m")
        month_end_date = required_date + pd.offsets.MonthEnd(0)

        return pd.Series({
            "week_start_date": week_start_date,
            "required_date": required_date,
            "plan_month": plan_month,
            "month_end_date": month_end_date
        })

    parsed_date = pd.to_datetime(value, errors="coerce")

    if not pd.isna(parsed_date):
        week_start_date = parsed_date - pd.to_timedelta(parsed_date.weekday(), unit="D")

        return pd.Series({
            "week_start_date": week_start_date,
            "required_date": parsed_date,
            "plan_month": parsed_date.strftime("%Y-%m"),
            "month_end_date": parsed_date + pd.offsets.MonthEnd(0)
        })

    return pd.Series({
        "week_start_date": pd.NaT,
        "required_date": pd.NaT,
        "plan_month": value,
        "month_end_date": pd.NaT
    })


def calculate_recommended_order_qty(shortage_qty, moq):
    """
    MOQ 기준 추천발주수량 계산

    원칙:
    1. 부족수량이 0이면 추천발주수량도 0
    2. MOQ가 0이면 부족수량 그대로 추천
    3. MOQ가 있으면 MOQ 배수로 올림 처리
    """
    if shortage_qty <= 0:
        return 0

    if moq <= 0:
        return shortage_qty

    return np.ceil(shortage_qty / moq) * moq


def adjust_excel_column_width(writer, sheet_name, df):
    """
    기존 3개 시트 공통 서식 적용 함수

    적용 내용:
    1. 모든 컬럼 너비 자동 조정
    2. amount 컬럼은 1,000 형식 적용
    3. date 컬럼은 yyyy-mm-dd 형식 적용
    4. risk_level 컬럼은 High / Medium / Low별 색상 적용
    """

    from openpyxl.styles import PatternFill, Font, Alignment

    worksheet = writer.sheets[sheet_name]

    high_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
    medium_fill = PatternFill(fill_type="solid", fgColor="FFEB9C")
    low_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
    normal_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")

    high_font = Font(color="9C0006")
    medium_font = Font(color="9C6500")
    low_font = Font(color="006100")
    normal_font = Font(color="000000")

    for idx, col in enumerate(df.columns, start=1):
        if len(df) > 0:
            max_length = max(
                df[col].astype(str).map(len).max(),
                len(str(col))
            )
        else:
            max_length = len(str(col))

        adjusted_width = min(max_length + 2, 35)
        column_letter = worksheet.cell(row=1, column=idx).column_letter
        worksheet.column_dimensions[column_letter].width = adjusted_width

        col_lower = str(col).lower()

        if "amount" in col_lower:
            for cell in worksheet[column_letter][1:]:
                cell.number_format = '#,##0'

        if "date" in col_lower:
            for cell in worksheet[column_letter][1:]:
                cell.number_format = 'yyyy-mm-dd'

        if col_lower == "risk_level":
            for cell in worksheet[column_letter][1:]:
                risk_value = str(cell.value).strip()

                if risk_value == "High":
                    cell.fill = high_fill
                    cell.font = high_font
                elif risk_value == "Medium":
                    cell.fill = medium_fill
                    cell.font = medium_font
                elif risk_value == "Low":
                    cell.fill = low_fill
                    cell.font = low_font
                else:
                    cell.fill = normal_fill
                    cell.font = normal_font

                cell.alignment = Alignment(horizontal="center")


def create_weekly_supply_view(
    supply_plan,
    required_by_material,
    po_open_valid,
    production_plan
):
    """
    weekly_supply_view 시트 생성 함수

    목적:
    자재별로 12주치 생산계획, 입고계획, 과부족 계산을 가로형으로 보여준다.

    구조:
    자재 1개당 3행 생성
    1) 생산계획
    2) 입고계획
    3) 과부족 계산

    과부족 계산 로직:
    첫 번째 주차:
        첫 주차 과부족 = 첫 주차 생산계획 - 현재고 - 첫 주차 입고계획

    두 번째 주차부터:
        이번 주차 과부족 = 이전 주차 과부족 + 이번 주차 생산계획 - 이번 주차 입고계획

    Excel 수식 예:
    생산계획 행: 2행
    입고계획 행: 3행
    과부족 계산 행: 4행
    current_stock: E열
    W22: G열
    W23: H열

    첫 주차:
        =G2-$E2-G3

    두 번째 주차:
        =G4+H2-H3

    세 번째 주차:
        =H4+I2-I3
    """

    from openpyxl.utils import get_column_letter

    valid_week_dates = production_plan["week_start_date"].dropna()

    if len(valid_week_dates) == 0:
        raise ValueError(
            "production_plan의 week_start_date가 모두 비어 있습니다. "
            "plan_week 날짜 변환 로직을 확인하세요."
        )

    start_week_date = pd.to_datetime(valid_week_dates.min())

    week_start_dates = [
        start_week_date + pd.Timedelta(weeks=i)
        for i in range(12)
    ]

    week_cols = [
        f"W{week_date.isocalendar().week:02d}"
        for week_date in week_start_dates
    ]

    required_weekly = required_by_material.copy()
    required_weekly["week_start_date"] = pd.to_datetime(
        required_weekly["week_start_date"],
        errors="coerce"
    )

    required_weekly = (
        required_weekly
        .groupby(
            ["material_code", "material_name", "week_start_date"],
            as_index=False
        )
        .agg(production_required_qty=("required_qty", "sum"))
    )

    production_lookup = {
        (
            row["material_code"],
            row["week_start_date"]
        ): row["production_required_qty"]
        for _, row in required_weekly.iterrows()
    }

    incoming_weekly = po_open_valid.copy()
    incoming_weekly["eta_date"] = pd.to_datetime(
        incoming_weekly["eta_date"],
        errors="coerce"
    )

    incoming_weekly["eta_week_start"] = (
        incoming_weekly["eta_date"]
        - pd.to_timedelta(incoming_weekly["eta_date"].dt.weekday, unit="D")
    )

    incoming_weekly = (
        incoming_weekly
        .groupby(
            ["material_code", "eta_week_start"],
            as_index=False
        )
        .agg(incoming_qty=("open_qty", "sum"))
    )

    incoming_lookup = {
        (
            row["material_code"],
            row["eta_week_start"]
        ): row["incoming_qty"]
        for _, row in incoming_weekly.iterrows()
    }

    material_base = (
        supply_plan[
            [
                "material_code",
                "material_name",
                "supplier",
                "lead_time_days",
                "current_stock"
            ]
        ]
        .drop_duplicates(subset=["material_code"])
        .sort_values("material_code")
        .reset_index(drop=True)
    )

    rows = []
    shortage_events = []
    shortage_value_lookup = {}

    fixed_columns_count = 6
    first_week_excel_col_idx = fixed_columns_count + 1

    for _, material in material_base.iterrows():
        material_code = material["material_code"]
        material_name = material["material_name"]
        supplier = material["supplier"]
        lead_time_days = material["lead_time_days"]
        current_stock = material["current_stock"]

        production_excel_row = len(rows) + 2
        incoming_excel_row = production_excel_row + 1
        shortage_excel_row = production_excel_row + 2

        production_row = {
            "material_code": material_code,
            "material_name": material_name,
            "supplier": supplier,
            "lead_time_days": lead_time_days,
            "current_stock": current_stock,
            "구분": "생산계획"
        }

        incoming_row = {
            "material_code": material_code,
            "material_name": material_name,
            "supplier": supplier,
            "lead_time_days": lead_time_days,
            "current_stock": current_stock,
            "구분": "입고계획"
        }

        shortage_row = {
            "material_code": material_code,
            "material_name": material_name,
            "supplier": supplier,
            "lead_time_days": lead_time_days,
            "current_stock": current_stock,
            "구분": "과부족 계산"
        }

        previous_shortage_value = None

        for week_idx, week_start_date in enumerate(week_start_dates):
            week_col = week_cols[week_idx]
            excel_col_letter = get_column_letter(first_week_excel_col_idx + week_idx)

            production_qty = production_lookup.get(
                (material_code, week_start_date),
                0
            )

            incoming_qty = incoming_lookup.get(
                (material_code, week_start_date),
                0
            )

            production_row[week_col] = production_qty
            incoming_row[week_col] = incoming_qty

            if week_idx == 0:
                shortage_value = production_qty - current_stock - incoming_qty

                shortage_formula = (
                    f"={excel_col_letter}{production_excel_row}"
                    f"-$E{production_excel_row}"
                    f"-{excel_col_letter}{incoming_excel_row}"
                )

            else:
                previous_excel_col_letter = get_column_letter(
                    first_week_excel_col_idx + week_idx - 1
                )

                shortage_value = previous_shortage_value + production_qty - incoming_qty

                shortage_formula = (
                    f"={previous_excel_col_letter}{shortage_excel_row}"
                    f"+{excel_col_letter}{production_excel_row}"
                    f"-{excel_col_letter}{incoming_excel_row}"
                )

            shortage_row[week_col] = shortage_formula
            previous_shortage_value = shortage_value

            shortage_value_lookup[(material_code, week_col)] = shortage_value

            if shortage_value > 0:
                shortage_events.append({
                    "material_code": material_code,
                    "material_name": material_name,
                    "supplier": supplier,
                    "week": week_col,
                    "week_start_date": week_start_date,
                    "shortage_qty": shortage_value
                })

        rows.append(production_row)
        rows.append(incoming_row)
        rows.append(shortage_row)

    weekly_supply_view = pd.DataFrame(rows)

    final_columns = [
        "material_code",
        "material_name",
        "supplier",
        "lead_time_days",
        "current_stock",
        "구분"
    ] + week_cols

    weekly_supply_view = weekly_supply_view[final_columns]
    shortage_events_df = pd.DataFrame(shortage_events)

    return weekly_supply_view, week_cols, shortage_events_df, shortage_value_lookup


def format_weekly_supply_view(writer, sheet_name, weekly_supply_view, week_cols):
    """
    weekly_supply_view 시트 전용 서식 적용 함수

    적용 내용:
    1. 첫 행 필터 적용
    2. 첫 행 고정
    3. 주차 컬럼 너비 조정
    4. 숫자 천 단위 구분 표시
    5. 구분 컬럼의 생산계획, 입고계획, 과부족 계산 행 배경색 적용
    6. 과부족 계산 행 굵은 글씨 적용
    7. 과부족 계산 행에서 양수 셀만 빨간 배경 조건부 서식 적용
    8. 자재별 3행 묶음 사이에 얇은 구분선 적용
    """

    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.utils import get_column_letter

    worksheet = writer.sheets[sheet_name]

    max_row = worksheet.max_row
    max_col = worksheet.max_column

    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.freeze_panes = "A2"

    production_fill = PatternFill(fill_type="solid", fgColor="DDEBF7")
    incoming_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    shortage_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")

    shortage_positive_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
    shortage_positive_font = Font(color="9C0006", bold=True)

    header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    header_font = Font(bold=True)

    thin_gray_side = Side(style="thin", color="D9D9D9")
    group_border = Border(bottom=thin_gray_side)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        header_value = worksheet.cell(row=1, column=col_idx).value

        if header_value in week_cols:
            worksheet.column_dimensions[col_letter].width = 12
        elif header_value == "material_name":
            worksheet.column_dimensions[col_letter].width = 22
        elif header_value == "supplier":
            worksheet.column_dimensions[col_letter].width = 16
        elif header_value == "구분":
            worksheet.column_dimensions[col_letter].width = 14
        else:
            worksheet.column_dimensions[col_letter].width = 15

        if header_value in week_cols or header_value in ["lead_time_days", "current_stock"]:
            for cell in worksheet[col_letter][1:]:
                cell.number_format = '#,##0'

    type_col_idx = list(weekly_supply_view.columns).index("구분") + 1
    first_week_col_idx = list(weekly_supply_view.columns).index(week_cols[0]) + 1
    last_week_col_idx = list(weekly_supply_view.columns).index(week_cols[-1]) + 1

    first_week_col_letter = get_column_letter(first_week_col_idx)
    last_week_col_letter = get_column_letter(last_week_col_idx)

    for row_idx in range(2, max_row + 1):
        row_type = worksheet.cell(row=row_idx, column=type_col_idx).value
        type_cell = worksheet.cell(row=row_idx, column=type_col_idx)

        if row_type == "생산계획":
            type_cell.fill = production_fill

        elif row_type == "입고계획":
            type_cell.fill = incoming_fill

        elif row_type == "과부족 계산":
            type_cell.fill = shortage_fill

            for col_idx in range(1, max_col + 1):
                worksheet.cell(row=row_idx, column=col_idx).font = Font(bold=True)

            shortage_range = (
                f"{first_week_col_letter}{row_idx}:"
                f"{last_week_col_letter}{row_idx}"
            )

            worksheet.conditional_formatting.add(
                shortage_range,
                CellIsRule(
                    operator="greaterThan",
                    formula=["0"],
                    fill=shortage_positive_fill,
                    font=shortage_positive_font
                )
            )

        type_cell.alignment = Alignment(horizontal="center")

        if (row_idx - 1) % 3 == 0:
            for col_idx in range(1, max_col + 1):
                worksheet.cell(row=row_idx, column=col_idx).border = group_border


# =========================================================
# 2. 파일 존재 여부 검증
# =========================================================

check_file_exists(production_plan_path, "production_plan.xlsx")
check_file_exists(bom_master_path, "bom_master.xlsx")
check_file_exists(inventory_path, "inventory.xlsx")
check_file_exists(po_open_path, "po_open.xlsx")
check_file_exists(material_master_path, "material_master.xlsx")

print("파일 존재 여부 검증 완료")


# =========================================================
# 3. 파일 읽기
# =========================================================

production_plan = pd.read_excel(production_plan_path)
bom_master = pd.read_excel(bom_master_path)
inventory = pd.read_excel(inventory_path)
po_open = pd.read_excel(po_open_path)
material_master = pd.read_excel(material_master_path)

production_plan = clean_columns(production_plan)
bom_master = clean_columns(bom_master)
inventory = clean_columns(inventory)
po_open = clean_columns(po_open)
material_master = clean_columns(material_master)

print_df_info("production_plan 원본", production_plan)
print_df_info("bom_master 원본", bom_master)
print_df_info("inventory 원본", inventory)
print_df_info("po_open 원본", po_open)
print_df_info("material_master 원본", material_master)


# =========================================================
# 4. 필수 컬럼 검증
# =========================================================

production_plan_required_cols = [
    "plan_week",
    "customer",
    "model",
    "product_code",
    "qty"
]

bom_master_required_cols = [
    "product_code",
    "material_code",
    "material_name",
    "usage_qty"
]

inventory_required_cols = [
    "material_code",
    "current_stock"
]

po_open_required_cols = [
    "material_code",
    "po_no",
    "eta_date",
    "open_qty"
]

material_master_required_cols = [
    "material_code",
    "supplier",
    "lead_time_days",
    "moq",
    "unit_price"
]

check_required_columns(production_plan, production_plan_required_cols, "production_plan.xlsx")
check_required_columns(bom_master, bom_master_required_cols, "bom_master.xlsx")
check_required_columns(inventory, inventory_required_cols, "inventory.xlsx")
check_required_columns(po_open, po_open_required_cols, "po_open.xlsx")
check_required_columns(material_master, material_master_required_cols, "material_master.xlsx")

print("\n필수 컬럼 검증 완료")


# =========================================================
# 5. 숫자 컬럼 처리
# =========================================================

production_plan = convert_numeric(production_plan, ["qty"])
bom_master = convert_numeric(bom_master, ["usage_qty"])
inventory = convert_numeric(inventory, ["current_stock"])
po_open = convert_numeric(po_open, ["open_qty"])
material_master = convert_numeric(material_master, ["lead_time_days", "moq", "unit_price"])

print("\n숫자 컬럼 변환 완료")


# =========================================================
# 6. 코드 컬럼 타입 정리
# =========================================================

production_plan = clean_code_column(production_plan, "product_code")
bom_master = clean_code_column(bom_master, "product_code")
bom_master = clean_code_column(bom_master, "material_code")
inventory = clean_code_column(inventory, "material_code")
po_open = clean_code_column(po_open, "material_code")
material_master = clean_code_column(material_master, "material_code")

print("\n코드 컬럼 타입 정리 완료")


# =========================================================
# 7. plan_week 날짜 정보 생성
# =========================================================

date_info = production_plan["plan_week"].apply(parse_iso_week)
production_plan = pd.concat([production_plan, date_info], axis=1)

print("\nplan_week 날짜 변환 결과")
print(
    production_plan[
        ["plan_week", "week_start_date", "required_date", "plan_month", "month_end_date"]
    ].drop_duplicates()
)

failed_date_count = production_plan["required_date"].isna().sum()

if failed_date_count > 0:
    print(f"\n경고: plan_week 날짜 변환 실패 행이 {failed_date_count:,}건 있습니다.")
    print("해당 행은 plan_month는 원본 문자열 기준으로 유지되며, 월별 입고 계산에서는 입고수량이 0으로 처리될 수 있습니다.")


# =========================================================
# 8. po_open eta_date 날짜 변환
# =========================================================

po_open["eta_date"] = pd.to_datetime(po_open["eta_date"], errors="coerce")

invalid_eta_count = po_open["eta_date"].isna().sum()

if invalid_eta_count > 0:
    print(f"\n경고: po_open.xlsx의 eta_date 중 날짜 변환 실패 행이 {invalid_eta_count:,}건 있습니다.")
    print("해당 행은 월별 입고예정수량 계산에서 제외됩니다.")

po_open_valid = po_open.dropna(subset=["eta_date"]).copy()

print_df_info("po_open 날짜 변환 후 유효 데이터", po_open_valid)


# =========================================================
# 9. BOM 누락 product_code 경고
# =========================================================

plan_product_codes = set(production_plan["product_code"].dropna().astype(str))
bom_product_codes = set(bom_master["product_code"].dropna().astype(str))

missing_bom_product_codes = sorted(list(plan_product_codes - bom_product_codes))

if missing_bom_product_codes:
    print("\n경고: BOM에 없는 product_code가 있습니다.")
    print("아래 product_code는 BOM이 없어 필요수량 계산에서 제외됩니다.")
    print(missing_bom_product_codes)
else:
    print("\nBOM 누락 product_code 없음")


# =========================================================
# 10. production_plan과 bom_master 병합
# =========================================================

plan_bom = production_plan.merge(
    bom_master,
    on="product_code",
    how="left",
    indicator=True
)

print_df_info("production_plan + bom_master 병합 결과", plan_bom)

bom_missing_rows = plan_bom[plan_bom["_merge"] == "left_only"]

if len(bom_missing_rows) > 0:
    print(f"\n경고: BOM 매칭 실패 행 수: {len(bom_missing_rows):,}")
    print("BOM이 없는 생산계획 행은 필요수량 계산에서 제외됩니다.")
    print(
        bom_missing_rows[
            ["plan_week", "customer", "model", "product_code", "qty"]
        ].drop_duplicates()
    )

plan_bom = plan_bom[plan_bom["_merge"] == "both"].copy()
plan_bom = plan_bom.drop(columns=["_merge"])


# =========================================================
# 11. 자재 필요수량 계산
# =========================================================

plan_bom["required_qty"] = plan_bom["qty"] * plan_bom["usage_qty"]

required_by_material = (
    plan_bom
    .groupby(
        [
            "plan_week",
            "plan_month",
            "month_end_date",
            "material_code",
            "material_name"
        ],
        as_index=False,
        dropna=False
    )
    .agg(
        required_qty=("required_qty", "sum"),
        week_start_date=("week_start_date", "min"),
        required_date=("required_date", "min")
    )
)

print_df_info("자재별 주차/월별 필요수량 required_by_material", required_by_material)
print(required_by_material.head())

if len(required_by_material) == 0:
    raise ValueError(
        "자재별 필요수량 required_by_material이 0행입니다. "
        "BOM merge, plan_week 변환, groupby 기준을 확인하세요."
    )


# =========================================================
# 12. 월별 입고예정수량 계산
# =========================================================

incoming_base = required_by_material[
    ["plan_week", "plan_month", "month_end_date", "material_code"]
].drop_duplicates()

incoming_merge = incoming_base.merge(
    po_open_valid[["material_code", "po_no", "eta_date", "open_qty"]],
    on="material_code",
    how="left"
)

incoming_merge["is_available_in_month"] = (
    incoming_merge["eta_date"].notna()
    & incoming_merge["month_end_date"].notna()
    & (incoming_merge["eta_date"] <= incoming_merge["month_end_date"])
)

incoming_merge["incoming_qty_for_month"] = np.where(
    incoming_merge["is_available_in_month"],
    incoming_merge["open_qty"],
    0
)

incoming_by_material_month = (
    incoming_merge
    .groupby(
        ["plan_week", "plan_month", "material_code"],
        as_index=False
    )
    .agg(incoming_qty=("incoming_qty_for_month", "sum"))
)

print_df_info("월별 입고예정수량 incoming_by_material_month", incoming_by_material_month)
print(incoming_by_material_month.head())


# =========================================================
# 13. 재고 정보 정리
# =========================================================

inventory_by_material = (
    inventory
    .groupby("material_code", as_index=False)
    .agg(current_stock=("current_stock", "sum"))
)

print_df_info("자재별 현재고 inventory_by_material", inventory_by_material)
print(inventory_by_material.head())


# =========================================================
# 14. 자재마스터 정리
# =========================================================

material_master_clean = (
    material_master
    .sort_values("material_code")
    .drop_duplicates(subset=["material_code"], keep="first")
    [
        [
            "material_code",
            "supplier",
            "lead_time_days",
            "moq",
            "unit_price"
        ]
    ]
)

print_df_info("자재마스터 정리 material_master_clean", material_master_clean)
print(material_master_clean.head())


# =========================================================
# 15. 필요수량 + 재고 + 월별 입고 + 자재마스터 연결
# =========================================================

supply_plan = required_by_material.copy()

supply_plan = supply_plan.merge(
    inventory_by_material,
    on="material_code",
    how="left"
)

supply_plan = supply_plan.merge(
    incoming_by_material_month,
    on=["plan_week", "plan_month", "material_code"],
    how="left"
)

supply_plan = supply_plan.merge(
    material_master_clean,
    on="material_code",
    how="left"
)

print_df_info("수급계획 병합 결과 supply_plan", supply_plan)

if len(supply_plan) == 0:
    raise ValueError(
        "최종 supply_plan이 0행입니다. "
        "수급계획 결과가 비어 있으므로 정상 산출로 볼 수 없습니다."
    )


# =========================================================
# 16. 빈값 처리
# =========================================================

supply_plan["current_stock"] = supply_plan["current_stock"].fillna(0)
supply_plan["incoming_qty"] = supply_plan["incoming_qty"].fillna(0)
supply_plan["supplier"] = supply_plan["supplier"].fillna("Unknown")
supply_plan["lead_time_days"] = supply_plan["lead_time_days"].fillna(0)
supply_plan["moq"] = supply_plan["moq"].fillna(0)
supply_plan["unit_price"] = supply_plan["unit_price"].fillna(0)

numeric_cols = [
    "required_qty",
    "current_stock",
    "incoming_qty",
    "lead_time_days",
    "moq",
    "unit_price"
]

supply_plan = convert_numeric(supply_plan, numeric_cols)


# =========================================================
# 17. 가용수량, 부족수량, 과잉수량 계산
# =========================================================

supply_plan["available_qty"] = (
    supply_plan["current_stock"] + supply_plan["incoming_qty"]
)

supply_plan["shortage_qty"] = (
    supply_plan["required_qty"] - supply_plan["available_qty"]
).clip(lower=0)

supply_plan["excess_qty"] = (
    supply_plan["available_qty"] - supply_plan["required_qty"]
).clip(lower=0)


# =========================================================
# 18. 부족금액, 과잉금액 계산
# =========================================================

supply_plan["shortage_amount"] = (
    supply_plan["shortage_qty"] * supply_plan["unit_price"]
)

supply_plan["excess_amount"] = (
    supply_plan["excess_qty"] * supply_plan["unit_price"]
)


# =========================================================
# 19. MOQ 기준 추천발주수량 계산
# =========================================================

supply_plan["recommended_order_qty"] = supply_plan.apply(
    lambda row: calculate_recommended_order_qty(
        row["shortage_qty"],
        row["moq"]
    ),
    axis=1
)

supply_plan["recommended_order_amount"] = (
    supply_plan["recommended_order_qty"] * supply_plan["unit_price"]
)

print("\nMOQ 기준 추천발주수량 계산 완료")
print(
    supply_plan[
        [
            "material_code",
            "shortage_qty",
            "moq",
            "recommended_order_qty",
            "recommended_order_amount"
        ]
    ].head(10)
)


# =========================================================
# 20. 리드타임 기준 발주필요일 계산
# =========================================================

supply_plan["order_required_date"] = supply_plan["required_date"] - pd.to_timedelta(
    supply_plan["lead_time_days"],
    unit="D"
)

print("\n리드타임 기준 발주필요일 계산 완료")
print(
    supply_plan[
        [
            "material_code",
            "required_date",
            "lead_time_days",
            "order_required_date"
        ]
    ].head(10)
)


# =========================================================
# 21. 리스크 등급 계산
# =========================================================

conditions = [
    (supply_plan["shortage_qty"] > 0) & (supply_plan["lead_time_days"] >= 45),
    (supply_plan["shortage_qty"] > 0) & (supply_plan["lead_time_days"] < 45),
    (supply_plan["shortage_qty"] == 0) & (supply_plan["excess_qty"] > 0),
    (supply_plan["shortage_qty"] == 0) & (supply_plan["excess_qty"] == 0)
]

choices = ["High", "Medium", "Low", "Normal"]

supply_plan["risk_level"] = np.select(
    conditions,
    choices,
    default="Normal"
)


# =========================================================
# 22. 계산 결과 검증
# =========================================================

validation_errors = []

if not (supply_plan["shortage_qty"] >= 0).all():
    validation_errors.append("shortage_qty에 음수가 있습니다.")

if not (supply_plan["excess_qty"] >= 0).all():
    validation_errors.append("excess_qty에 음수가 있습니다.")

if not np.allclose(
    supply_plan["available_qty"],
    supply_plan["current_stock"] + supply_plan["incoming_qty"]
):
    validation_errors.append("available_qty 계산이 current_stock + incoming_qty와 일치하지 않습니다.")

if not np.allclose(
    supply_plan["shortage_amount"],
    supply_plan["shortage_qty"] * supply_plan["unit_price"]
):
    validation_errors.append("shortage_amount 계산이 shortage_qty × unit_price와 일치하지 않습니다.")

if not np.allclose(
    supply_plan["excess_amount"],
    supply_plan["excess_qty"] * supply_plan["unit_price"]
):
    validation_errors.append("excess_amount 계산이 excess_qty × unit_price와 일치하지 않습니다.")

if not np.allclose(
    supply_plan["recommended_order_amount"],
    supply_plan["recommended_order_qty"] * supply_plan["unit_price"]
):
    validation_errors.append("recommended_order_amount 계산이 recommended_order_qty × unit_price와 일치하지 않습니다.")

if (supply_plan["recommended_order_qty"] < supply_plan["shortage_qty"]).any():
    validation_errors.append("recommended_order_qty가 shortage_qty보다 작은 행이 있습니다.")

if validation_errors:
    print("\n계산 결과 검증 실패")
    for error in validation_errors:
        print(f"- {error}")

    raise ValueError("계산 결과 검증 중 오류가 발견되었습니다.")
else:
    print("\n계산 결과 검증 완료")


# =========================================================
# 23. 최종 컬럼 정리
# =========================================================

final_columns = [
    "plan_week",
    "plan_month",
    "week_start_date",
    "required_date",
    "month_end_date",
    "material_code",
    "material_name",
    "required_qty",
    "current_stock",
    "incoming_qty",
    "available_qty",
    "shortage_qty",
    "shortage_amount",
    "excess_qty",
    "excess_amount",
    "supplier",
    "lead_time_days",
    "moq",
    "unit_price",
    "recommended_order_qty",
    "recommended_order_amount",
    "order_required_date",
    "risk_level"
]

supply_plan = supply_plan[final_columns]


# =========================================================
# 24. 정렬
# =========================================================

risk_order = {
    "High": 1,
    "Medium": 2,
    "Low": 3,
    "Normal": 4
}

supply_plan["risk_sort"] = supply_plan["risk_level"].map(risk_order)

supply_plan = (
    supply_plan
    .sort_values(
        by=["risk_sort", "shortage_amount", "material_code"],
        ascending=[True, False, True]
    )
    .drop(columns=["risk_sort"])
    .reset_index(drop=True)
)

print_df_info("최종 수급계획 supply_plan", supply_plan)
print(supply_plan.head(10))


# =========================================================
# 25. 요약 시트 생성: summary_by_risk
# =========================================================

summary_by_risk = (
    supply_plan
    .groupby("risk_level", as_index=False)
    .agg(
        material_count=("material_code", "nunique"),
        total_required_qty=("required_qty", "sum"),
        total_shortage_qty=("shortage_qty", "sum"),
        total_shortage_amount=("shortage_amount", "sum"),
        total_excess_qty=("excess_qty", "sum"),
        total_excess_amount=("excess_amount", "sum"),
        total_recommended_order_qty=("recommended_order_qty", "sum"),
        total_recommended_order_amount=("recommended_order_amount", "sum")
    )
)

summary_by_risk["risk_sort"] = summary_by_risk["risk_level"].map(risk_order)

summary_by_risk = (
    summary_by_risk
    .sort_values("risk_sort")
    .drop(columns=["risk_sort"])
    .reset_index(drop=True)
)

print_df_info("리스크별 요약 summary_by_risk", summary_by_risk)
print(summary_by_risk)


# =========================================================
# 26. 요약 시트 생성: summary_by_supplier
# =========================================================

supplier_risk_count = (
    supply_plan
    .pivot_table(
        index="supplier",
        columns="risk_level",
        values="material_code",
        aggfunc="nunique",
        fill_value=0
    )
    .reset_index()
)

if "High" not in supplier_risk_count.columns:
    supplier_risk_count["High"] = 0

if "Medium" not in supplier_risk_count.columns:
    supplier_risk_count["Medium"] = 0

supplier_amount_summary = (
    supply_plan
    .groupby("supplier", as_index=False)
    .agg(
        material_count=("material_code", "nunique"),
        total_shortage_qty=("shortage_qty", "sum"),
        total_shortage_amount=("shortage_amount", "sum"),
        total_excess_qty=("excess_qty", "sum"),
        total_excess_amount=("excess_amount", "sum"),
        total_recommended_order_qty=("recommended_order_qty", "sum"),
        total_recommended_order_amount=("recommended_order_amount", "sum")
    )
)

summary_by_supplier = supplier_amount_summary.merge(
    supplier_risk_count[["supplier", "High", "Medium"]],
    on="supplier",
    how="left"
)

summary_by_supplier = summary_by_supplier.rename(
    columns={
        "High": "high_risk_count",
        "Medium": "medium_risk_count"
    }
)

summary_by_supplier["high_risk_count"] = summary_by_supplier["high_risk_count"].fillna(0).astype(int)
summary_by_supplier["medium_risk_count"] = summary_by_supplier["medium_risk_count"].fillna(0).astype(int)

summary_by_supplier = summary_by_supplier[
    [
        "supplier",
        "material_count",
        "high_risk_count",
        "medium_risk_count",
        "total_shortage_qty",
        "total_shortage_amount",
        "total_excess_qty",
        "total_excess_amount",
        "total_recommended_order_qty",
        "total_recommended_order_amount"
    ]
]

summary_by_supplier = summary_by_supplier.sort_values(
    by=["total_shortage_amount", "high_risk_count", "supplier"],
    ascending=[False, False, True]
).reset_index(drop=True)

print_df_info("협력사별 요약 summary_by_supplier", summary_by_supplier)
print(summary_by_supplier.head(10))


# =========================================================
# 27. weekly_supply_view 시트 생성
# =========================================================

weekly_supply_view, week_cols, weekly_shortage_events, shortage_value_lookup = create_weekly_supply_view(
    supply_plan=supply_plan,
    required_by_material=required_by_material,
    po_open_valid=po_open_valid,
    production_plan=production_plan
)

print_df_info("주차별 수급 현황 weekly_supply_view", weekly_supply_view)
print(weekly_supply_view.head(12))


# =========================================================
# 28. weekly_supply_view 검증
# =========================================================

weekly_validation_errors = []

weekly_material_count = weekly_supply_view["material_code"].nunique()
expected_weekly_rows = weekly_material_count * 3

if len(week_cols) != 12:
    weekly_validation_errors.append(
        f"주차 컬럼 수가 12개가 아닙니다. 현재: {len(week_cols)}개"
    )

if len(weekly_supply_view) != expected_weekly_rows:
    weekly_validation_errors.append(
        f"weekly_supply_view 행 수가 자재 수 × 3과 일치하지 않습니다. "
        f"현재 행 수: {len(weekly_supply_view)}, 예상 행 수: {expected_weekly_rows}"
    )

material_row_check = weekly_supply_view.groupby("material_code")["구분"].nunique()

if not (material_row_check == 3).all():
    weekly_validation_errors.append(
        "일부 자재가 생산계획/입고계획/과부족 계산 3행 구조를 갖지 않습니다."
    )

required_types = {"생산계획", "입고계획", "과부족 계산"}
actual_types = set(weekly_supply_view["구분"].unique())

if actual_types != required_types:
    weekly_validation_errors.append(
        f"구분 컬럼 값이 예상과 다릅니다. 현재 값: {actual_types}"
    )

for material_code in weekly_supply_view["material_code"].unique():
    material_rows = weekly_supply_view[weekly_supply_view["material_code"] == material_code]

    production_row = material_rows[material_rows["구분"] == "생산계획"]
    incoming_row = material_rows[material_rows["구분"] == "입고계획"]
    shortage_row = material_rows[material_rows["구분"] == "과부족 계산"]

    if len(production_row) != 1 or len(incoming_row) != 1 or len(shortage_row) != 1:
        weekly_validation_errors.append(
            f"{material_code}의 3행 구조가 올바르지 않습니다."
        )
        continue

    current_stock = float(production_row["current_stock"].iloc[0])
    previous_shortage = None

    for week_idx, week_col in enumerate(week_cols):
        production_qty = float(production_row[week_col].iloc[0])
        incoming_qty = float(incoming_row[week_col].iloc[0])
        calculated_shortage = float(shortage_value_lookup[(material_code, week_col)])

        if week_idx == 0:
            expected_shortage = production_qty - current_stock - incoming_qty
        else:
            expected_shortage = previous_shortage + production_qty - incoming_qty

        if not np.isclose(calculated_shortage, expected_shortage):
            weekly_validation_errors.append(
                f"{material_code} {week_col} 과부족 계산이 맞지 않습니다. "
                f"계산값: {calculated_shortage}, 예상값: {expected_shortage}"
            )

        previous_shortage = calculated_shortage

if weekly_validation_errors:
    print("\nweekly_supply_view 검증 실패")
    for error in weekly_validation_errors:
        print(f"- {error}")

    raise ValueError("weekly_supply_view 검증 중 오류가 발견되었습니다.")

else:
    print("\nweekly_supply_view 검증 완료")
    print(f"- 전체 자재 수: {weekly_material_count:,}")
    print(f"- 전체 주차 수: {len(week_cols):,}")
    print(f"- 전체 행 수: {len(weekly_supply_view):,}")


# =========================================================
# 29. 최종 결과 해석 가능 여부 검증
# =========================================================

interpretation_required_columns = [
    "plan_week",
    "plan_month",
    "material_code",
    "material_name",
    "required_qty",
    "current_stock",
    "incoming_qty",
    "available_qty",
    "shortage_qty",
    "shortage_amount",
    "excess_qty",
    "excess_amount",
    "supplier",
    "lead_time_days",
    "moq",
    "unit_price",
    "recommended_order_qty",
    "recommended_order_amount",
    "order_required_date",
    "risk_level"
]

check_required_columns(
    supply_plan,
    interpretation_required_columns,
    "최종 material_supply_plan"
)

if len(supply_plan) == 0:
    raise ValueError("최종 결과가 0행입니다. 정상적인 수급계획서로 볼 수 없습니다.")

if "weekly_supply_view" == "":
    raise ValueError("weekly_supply_view 시트명이 비어 있습니다.")

print("\n최종 결과 해석 가능 여부 검증 완료")


# =========================================================
# 30. Excel 저장
# =========================================================

with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    supply_plan.to_excel(
        writer,
        sheet_name="material_supply_plan",
        index=False
    )

    summary_by_risk.to_excel(
        writer,
        sheet_name="summary_by_risk",
        index=False
    )

    summary_by_supplier.to_excel(
        writer,
        sheet_name="summary_by_supplier",
        index=False
    )

    weekly_supply_view.to_excel(
        writer,
        sheet_name="weekly_supply_view",
        index=False
    )

    adjust_excel_column_width(writer, "material_supply_plan", supply_plan)
    adjust_excel_column_width(writer, "summary_by_risk", summary_by_risk)
    adjust_excel_column_width(writer, "summary_by_supplier", summary_by_supplier)

    format_weekly_supply_view(
        writer=writer,
        sheet_name="weekly_supply_view",
        weekly_supply_view=weekly_supply_view,
        week_cols=week_cols
    )

print("\nExcel 저장 완료")


# =========================================================
# 31. 실행 후 확인용 조회 코드
# =========================================================

print("\n결과 파일 생성 여부")
print(output_path.exists())

high_risk_count = (supply_plan["risk_level"] == "High").sum()
medium_risk_count = (supply_plan["risk_level"] == "Medium").sum()

total_shortage_qty = supply_plan["shortage_qty"].sum()
total_shortage_amount = supply_plan["shortage_amount"].sum()

total_excess_qty = supply_plan["excess_qty"].sum()
total_excess_amount = supply_plan["excess_amount"].sum()

total_recommended_order_qty = supply_plan["recommended_order_qty"].sum()
total_recommended_order_amount = supply_plan["recommended_order_amount"].sum()

if len(supply_plan) > 0 and supply_plan["shortage_amount"].max() > 0:
    top_shortage_material = supply_plan.sort_values(
        "shortage_amount",
        ascending=False
    ).head(1)
else:
    top_shortage_material = pd.DataFrame()

print("\n부족금액이 가장 큰 자재")
if len(top_shortage_material) > 0:
    print(
        top_shortage_material[
            [
                "plan_week",
                "plan_month",
                "required_date",
                "material_code",
                "material_name",
                "shortage_qty",
                "shortage_amount",
                "moq",
                "recommended_order_qty",
                "recommended_order_amount",
                "supplier",
                "lead_time_days",
                "order_required_date",
                "risk_level"
            ]
        ]
    )
else:
    print("부족금액이 있는 자재가 없습니다.")


# =========================================================
# 32. weekly_supply_view 부족 요약
# =========================================================

if len(weekly_shortage_events) > 0:
    weekly_shortage_material_count = weekly_shortage_events["material_code"].nunique()
    weekly_shortage_week_count = weekly_shortage_events["week"].nunique()

    first_shortage_week = (
        weekly_shortage_events
        .sort_values("week_start_date")
        .iloc[0]["week"]
    )

    max_shortage_row = (
        weekly_shortage_events
        .sort_values("shortage_qty", ascending=False)
        .iloc[0]
    )

    max_shortage_material = max_shortage_row["material_code"]
    max_shortage_qty = max_shortage_row["shortage_qty"]

    print("\n부족 발생 자재/주차 요약")
    print(
        weekly_shortage_events
        .sort_values(["week_start_date", "material_code"])
        .head(20)
    )

else:
    weekly_shortage_material_count = 0
    weekly_shortage_week_count = 0
    first_shortage_week = "부족 없음"
    max_shortage_material = "부족 없음"
    max_shortage_qty = 0

    print("\n부족 발생 자재/주차 요약")
    print("부족이 발생한 자재와 주차가 없습니다.")


# =========================================================
# 33. 완료 보고
# =========================================================

print("\n==============================")
print("자재 수급계획 생성 완료")
print("==============================")
print(f"생성 완료 파일 경로: {output_path}")
print(f"전체 행 수: {len(supply_plan):,}")
print(f"전체 자재 수: {supply_plan['material_code'].nunique():,}")
print(f"High 리스크 자재 수: {high_risk_count:,}")
print(f"Medium 리스크 자재 수: {medium_risk_count:,}")
print(f"총 부족수량: {total_shortage_qty:,.0f}")
print(f"총 부족금액: {total_shortage_amount:,.0f}")
print(f"총 과잉수량: {total_excess_qty:,.0f}")
print(f"총 과잉금액: {total_excess_amount:,.0f}")
print(f"총 추천발주수량: {total_recommended_order_qty:,.0f}")
print(f"총 추천발주금액: {total_recommended_order_amount:,.0f}")

print("\n==============================")
print("weekly_supply_view 생성 완료")
print("==============================")
print("weekly_supply_view 생성 완료 여부: True")
print(f"전체 자재 수: {weekly_material_count:,}")
print(f"전체 주차 수: {len(week_cols):,}")
print(f"부족 발생 자재 수: {weekly_shortage_material_count:,}")
print(f"부족 발생 주차 수: {weekly_shortage_week_count:,}")
print(f"가장 먼저 부족이 발생한 주차: {first_shortage_week}")
print(f"최대 부족수량이 발생한 자재: {max_shortage_material}")
print(f"최대 부족수량: {max_shortage_qty:,.0f}")


# =========================================================
# 34. 특정 material_code 검색 예시
# =========================================================

search_material_code = ""

if search_material_code:
    search_result = supply_plan[
        supply_plan["material_code"].astype(str) == str(search_material_code)
    ]

    print(f"\nmaterial_code 검색 결과: {search_material_code}")
    print(search_result)
